import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import Intervento

EXCEL_FILE = "data/interventi_emissioni.xlsx"
SHEET_NAME = "Interventi"

HEADERS = [
    "ID",
    "Tipologia Sito (ME)",
    "Tipologia Sito (LCA)",
    "Tipologia di Materiale",
    "Pressione Esercizio (bar)",
    "Classificazione Dispersione",
    "Tipologia Riparazione",
    "Interruzione Fornitura",
    "Data Rilevamento Perdita",
    "Data Esecuzione Riparazione",
    "Image Path",
    "Unità di Misura Emissione",
    "Valore Emissione",
    "PPM",
    "Kg/h CH4",
    "Fattore Emissione Kg/h CO2",
]

# Colori
COLOR_HEADER_BG = "2E4057"  # Blu scuro
COLOR_HEADER_FG = "FFFFFF"  # Bianco
COLOR_SURVEY_BG = "00B4D8"  # Teal (Survey)
COLOR_EMISSION_BG = "7B2FBE"  # Viola (Emissioni)
COLOR_ROW_ALT = "F0F4FF"  # Azzurro chiaro alternato
COLOR_RESULT_BG = "E8F5E9"  # Verde chiaro per risultati


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _create_workbook(filepath: str) -> Workbook:
    """Crea un nuovo workbook con intestazioni formattate"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Riga 1: categoria Survey vs Emissione
    survey_cols = 11
    emission_cols = 5

    # Merge celle categoria Survey
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=survey_cols)
    cell_s = ws.cell(row=1, column=1, value="SURVEY")
    cell_s.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    cell_s.fill = PatternFill("solid", start_color="0D9488")
    cell_s.alignment = Alignment(horizontal="center", vertical="center")

    # Merge celle categoria Emissione
    ws.merge_cells(
        start_row=1,
        start_column=survey_cols + 1,
        end_row=1,
        end_column=survey_cols + emission_cols,
    )
    cell_e = ws.cell(row=1, column=survey_cols + 1, value="EMISSIONE & CONVERSIONE")
    cell_e.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    cell_e.fill = PatternFill("solid", start_color="7B2FBE")
    cell_e.alignment = Alignment(horizontal="center", vertical="center")

    # Riga 2: headers colonne
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10, name="Arial")
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"

    # Larghezze colonne
    col_widths = [12, 22, 16, 20, 20, 25, 22, 18, 22, 24, 16, 16, 14, 14, 22, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)
    return wb


def save_intervento(
    intervento: Intervento, filepath: str = EXCEL_FILE, image_file: str = None
) -> int:
    """
    Salva un intervento nel file Excel.
    Se fornito, salva anche l'immagine in data/image/{image_id}/
    Ritorna il numero di riga scritto.
    """
    if not os.path.exists(filepath):
        _create_workbook(filepath)

    # Gestisci salvataggio immagine se fornita
    image_path = ""
    if image_file:
        # Genera ID univoco basato su timestamp
        image_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        intervento.survey.image_id = image_id

        # Crea cartella per l'immagine
        image_folder = f"data/image/{image_id}"
        os.makedirs(image_folder, exist_ok=True)

        # Copia l'immagine nella cartella
        try:
            filename = os.path.basename(image_file)
            image_path = os.path.join(image_folder, filename)
            shutil.copy2(image_file, image_path)
            intervento.survey.image_path = image_path
        except Exception as e:
            print(f"Errore salvataggio immagine: {e}")
            image_path = ""

    wb = load_workbook(filepath)
    ws = wb[SHEET_NAME]

    # Trova prima riga vuota (dopo le 2 di intestazione)
    row = ws.max_row + 1
    if row < 3:
        row = 3

    data = intervento.to_excel_row()
    is_alt = row % 2 == 0

    survey_fill = PatternFill("solid", start_color="E0F7FA" if is_alt else "FFFFFF")
    emission_fill = PatternFill("solid", start_color="F3E5FF" if is_alt else "FFFFFF")
    result_fill = PatternFill("solid", start_color="E8F5E9")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=data.get(header, ""))
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

        # Colorazione per categoria
        if col_idx <= 11:
            cell.fill = survey_fill
        elif col_idx <= 16:
            cell.fill = emission_fill
        else:
            cell.fill = result_fill
            cell.font = Font(name="Arial", size=10, bold=True, color="1B5E20")

        # Formato numerico
        if header in ("PPM", "Kg/h CH4", "Fattore Emissione Kg/h CO2"):
            cell.number_format = "#,##0.000000"

    ws.row_dimensions[row].height = 18
    wb.save(filepath)
    return row


def get_all_interventi(filepath: str = EXCEL_FILE) -> list[dict]:
    if not os.path.exists(filepath):
        return []
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb[SHEET_NAME]
        records = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if any(v for v in row if v is not None):
                records.append(dict(zip(HEADERS, row)))
        wb.close()
        return records
    except Exception as e:
        print(f"Errore lettura Excel: {e}")
        return []
